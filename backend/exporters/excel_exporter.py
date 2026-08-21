import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from ..models.schemas import CueSheetManifest

def export_cue_sheet_to_excel(manifest: CueSheetManifest) -> bytes:
    """
    Generates a broadcast-standard PMA / CISAC compliant Excel (.xlsx) Cue Sheet.
    Returns bytes buffer ready for download or disk write.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Music Cue Sheet"
    ws.views.sheetView[0].showGridLines = True

    # Palette Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    meta_label_font = Font(name="Calibri", size=10, bold=True, color="4B5563")
    meta_val_font = Font(name="Calibri", size=10, color="111827")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="111827")
    mono_font = Font(name="Consolas", size=10, color="1F2937")

    brand_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    verified_fill = PatternFill(start_color="DEF7EC", end_color="DEF7EC", fill_type="solid")
    flagged_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="E2E8F0")
    table_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # 1. Studio Header Banner (Rows 1-2)
    ws.merge_cells("A1:M2")
    banner_cell = ws["A1"]
    banner_cell.value = f"CUECLEAR BROADCAST MUSIC CUE SHEET: {manifest.project_title.upper()}"
    banner_cell.font = title_font
    banner_cell.fill = brand_fill
    banner_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # 2. Metadata Block (Rows 4-6)
    meta_items = [
        ("Production Company:", manifest.production_company, "Target Distributor:", manifest.target_distributor),
        ("Director:", manifest.director, "Total Audio Cues:", f"{manifest.total_cues} ({manifest.cleared_cues} Cleared)"),
        ("Framerate:", f"{manifest.framerate} FPS", "Compliance Health:", f"{manifest.compliance_score}% Cleared")
    ]

    for row_idx, (l1, v1, l2, v2) in enumerate(meta_items, start=4):
        # Col A-B
        ws.cell(row=row_idx, column=1, value=l1).font = meta_label_font
        ws.cell(row=row_idx, column=2, value=v1).font = meta_val_font
        # Col D-E
        ws.cell(row=row_idx, column=4, value=l2).font = meta_label_font
        ws.cell(row=row_idx, column=5, value=v2).font = meta_val_font

    # 3. Table Column Headers (Row 8)
    headers = [
        ("Cue #", 8, "center"),
        ("Cue Title", 26, "left"),
        ("Artist / Performer", 20, "left"),
        ("Usage", 8, "center"),
        ("Timecode In", 14, "center"),
        ("Timecode Out", 14, "center"),
        ("Duration", 12, "center"),
        ("Composer(s) & PRO", 30, "left"),
        ("Writer %", 14, "center"),
        ("Publisher(s) & PRO", 30, "left"),
        ("Pub %", 14, "center"),
        ("Work ID / ISWC", 24, "left"),
        ("Status", 22, "center")
    ]

    table_header_row = 8
    for col_idx, (h_text, _, align) in enumerate(headers, start=1):
        cell = ws.cell(row=table_header_row, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        cell.border = table_border
    ws.row_dimensions[table_header_row].height = 25

    # 4. Table Data Rows
    current_row = 9
    for idx, cue in enumerate(manifest.cues):
        writers_str = "\n".join([f"{w.name} ({w.pro})" for w in cue.writers]) or "N/A"
        writer_shares_str = "\n".join([f"{w.share:.0f}%" if w.share is not None else "Undisclosed" for w in cue.writers]) or "N/A"
        
        pubs_str = "\n".join([f"{p.name} ({p.pro})" for p in cue.publishers]) or "N/A"
        pub_shares_str = "\n".join([f"{p.share:.0f}%" if p.share is not None else "Undisclosed" for p in cue.publishers]) or "N/A"

        work_id_str = f"Work: {cue.work_id or 'N/A'}\nISWC: {cue.iswc or 'N/A'}"
        if cue.supervisor_signed_off:
            status_str = "CLEARED (SUPERVISOR SIGN-OFF)"
        elif cue.is_verified:
            status_str = "CLEARED"
        else:
            status_str = "ACTION REQUIRED (Undisclosed / Incomplete)"

        row_values = [
            (f"{cue.cue_number:02d}", mono_font, "center"),
            (cue.title, data_font, "left"),
            (cue.artist or "N/A", data_font, "left"),
            (cue.usage_type.value, mono_font, "center"),
            (cue.timecode_in, mono_font, "center"),
            (cue.timecode_out, mono_font, "center"),
            (cue.duration_timecode, mono_font, "center"),
            (writers_str, data_font, "left"),
            (writer_shares_str, mono_font, "center"),
            (pubs_str, data_font, "left"),
            (pub_shares_str, mono_font, "center"),
            (work_id_str, mono_font, "left"),
            (status_str, Font(name="Calibri", size=9, bold=True, color="065F46" if cue.is_verified else "991B1B"), "center")
        ]

        row_fill = verified_fill if cue.is_verified else flagged_fill
        if idx % 2 == 1 and cue.is_verified:
            row_fill = alt_row_fill

        for col_idx, (val, font_style, align) in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font_style
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = table_border
            if col_idx == 13:  # Status cell color
                cell.fill = verified_fill if cue.is_verified else flagged_fill
            elif row_fill:
                cell.fill = row_fill

        ws.row_dimensions[current_row].height = max(24, len(cue.writers) * 18, len(cue.publishers) * 18)
        current_row += 1

    # 5. Set column widths
    for col_idx, (_, col_width, _) in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width

    # Write to memory buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
