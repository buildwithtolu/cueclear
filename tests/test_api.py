import io
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.main import app, safe_filename_part, SESSIONS
from backend.models.schemas import (
    CueSheetManifest,
    ResolvedCue,
    RightsHolder,
    UsageType,
)
from backend.agent.split_reconciler import validate_splits


client = TestClient(app)


@pytest.fixture(autouse=True)
def _clear_sessions():
    SESSIONS.clear()
    yield
    SESSIONS.clear()


def test_health_route():
    res = client.get("/api/health")
    assert res.status_code == 200
    data = res.json()
    assert data["status"] == "healthy"
    # Local/dev mode exposes detail flags; public mode should not.
    if "parallel_integration" in data:
        assert data["parallel_integration"] == "search_and_extract"
        assert "parallel_search_configured" in data


def test_sample_timelines_route():
    res = client.get("/api/sample-timelines")
    assert res.status_code == 200
    samples = res.json()
    assert len(samples) >= 2
    assert samples[0]["id"] == "sample_trailer"


def test_upload_and_clearance_lifecycle(monkeypatch):
    monkeypatch.setenv("GEMINI_API_KEY", "")
    monkeypatch.setenv("GOOGLE_API_KEY", "")
    monkeypatch.setenv("PARALLEL_API_KEY", "")
    monkeypatch.setattr("backend.agent.parallel_tool.PARALLEL_API_KEY", "")
    monkeypatch.setattr("backend.main.agent.has_gemini", False)

    with open("samples/sample_trailer.edl", "rb") as edl_file:
        upload_res = client.post(
            "/api/upload-timeline",
            data={
                "file_type": "edl",
                "project_title": "Neon Mirage Master",
            },
            files={
                "file": ("sample_trailer.edl", edl_file, "text/plain"),
            },
        )
    assert upload_res.status_code == 200
    upload_data = upload_res.json()
    assert upload_data["total_clips"] == 5
    assert "cueclear_sid" in upload_res.cookies

    clearance_res = client.post(
        "/api/run-clearance",
        json={
            "project_title": "Neon Mirage Master",
            "clips": upload_data["clips"],
        },
    )
    assert clearance_res.status_code == 200
    manifest = clearance_res.json()
    assert manifest["total_cues"] == 5
    assert manifest["cleared_cues"] >= 2
    assert manifest["compliance_score"] >= 40.0

    excel_res = client.get("/api/export/excel")
    assert excel_res.status_code == 200
    assert "spreadsheetml" in excel_res.headers["content-type"]
    assert len(excel_res.content) > 1000

    xml_res = client.get("/api/export/cisac-xml")
    assert xml_res.status_code == 200
    assert "xml" in xml_res.headers["content-type"]
    assert b"AVCueSheet" in xml_res.content

    json_res = client.get("/api/export/json")
    assert json_res.status_code == 200
    json_data = json_res.json()
    assert json_data["project_title"] == "Neon Mirage Master"


def test_export_requires_session_manifest():
    fresh = TestClient(app)
    res = fresh.get("/api/export/excel")
    assert res.status_code == 400


def test_stream_requires_uploaded_timeline():
    fresh = TestClient(app)
    res = fresh.get("/api/stream-clearance")
    assert res.status_code == 400


def test_sessions_are_isolated(monkeypatch):
    monkeypatch.setenv("GEMINI_API_KEY", "")
    monkeypatch.setenv("GOOGLE_API_KEY", "")
    monkeypatch.setenv("PARALLEL_API_KEY", "")
    monkeypatch.setattr("backend.agent.parallel_tool.PARALLEL_API_KEY", "")
    monkeypatch.setattr("backend.main.agent.has_gemini", False)

    client_a = TestClient(app)
    client_b = TestClient(app)

    with open("samples/sample_mixed_clearance.edl", "rb") as edl_file:
        up_a = client_a.post(
            "/api/upload-timeline",
            data={"file_type": "edl", "project_title": "Session A"},
            files={"file": ("mixed.edl", edl_file, "text/plain")},
        )
    assert up_a.status_code == 200
    clips_a = up_a.json()["clips"]

    run_a = client_a.post(
        "/api/run-clearance",
        json={"project_title": "Session A", "clips": clips_a},
    )
    assert run_a.status_code == 200

    # Client B has no shared manifest
    export_b = client_b.get("/api/export/excel")
    assert export_b.status_code == 400

    export_a = client_a.get("/api/export/excel")
    assert export_a.status_code == 200


def test_frontend_static_serving():
    res = client.get("/")
    assert res.status_code == 200
    assert "CueClear" in res.text
    assert "Upload timeline" in res.text
    assert "Theses" not in res.text
    assert "HOLLYWOOD" not in res.text
    assert "IDENTITY ARCHETYPES" not in res.text
    assert "CUE" in res.text
    assert "CLEAR" in res.text
    assert 'rel="icon"' in res.text
    assert "/favicon.svg" in res.text


def test_sign_off_persists_into_export_manifest():
    pending = ResolvedCue(
        cue_number=7,
        title="Exit Music (For A Film)",
        artist="Radiohead",
        usage_type=UsageType.BI,
        timecode_in="00:00:00:00",
        timecode_out="00:01:00:00",
        duration_frames=1440,
        duration_timecode="00:01:00:00",
        writers=[
            RightsHolder(name="Yorke", share=None, pro="BMI"),
            RightsHolder(name="Greenwood", share=None, pro="BMI"),
        ],
        publishers=[RightsHolder(name="Warner Chappell", share=None, pro="BMI")],
    )
    validate_splits(pending)
    assert pending.is_verified is False

    # Establish session cookie, then inject manifest into that session
    bootstrap = client.post(
        "/api/upload-timeline",
        data={
            "raw_content": "TITLE: X\n001  A1       NONE  C        00:00:00:00 00:00:01:00 00:00:00:00 00:00:01:00\n* FROM CLIP NAME: Temp.wav\n",
            "file_type": "edl",
            "project_title": "SignOff Export Test",
        },
    )
    assert bootstrap.status_code == 200
    sid = bootstrap.cookies.get("cueclear_sid")
    assert sid
    assert sid in SESSIONS
    SESSIONS[sid].manifest = CueSheetManifest(
        project_title="SignOff Export Test",
        cues=[pending],
        total_cues=1,
        cleared_cues=0,
        flagged_cues=1,
        compliance_score=0.0,
    )

    sign_res = client.post(
        "/api/sign-off",
        json={"cue_number": 7, "signed_off_by": "Music Supervisor"},
    )
    assert sign_res.status_code == 200
    body = sign_res.json()
    assert body["cleared_cues"] == 1
    assert body["compliance_score"] == 100.0
    assert body["cues"][0]["supervisor_signed_off"] is True
    assert body["cues"][0]["is_verified"] is True

    excel_res = client.get("/api/export/excel")
    assert excel_res.status_code == 200
    wb = load_workbook(io.BytesIO(excel_res.content))
    ws = wb.active
    status_values = [
        cell.value
        for row in ws.iter_rows(min_row=9, max_col=13)
        for cell in row
        if cell.column == 13
    ]
    assert any(isinstance(v, str) and "SUPERVISOR SIGN-OFF" in v for v in status_values)


def test_safe_filename_part_strips_injection():
    assert ".." not in safe_filename_part('../../evil\r\nX')
    assert '"' not in safe_filename_part('a"b')
    assert safe_filename_part("My Cool Project!") == "My_Cool_Project"
