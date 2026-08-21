import pytest
import io
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

from backend.models.schemas import (
    CueSheetManifest,
    ResolvedCue,
    RightsHolder,
    UsageType,
    SplitStatus
)
from backend.exporters.excel_exporter import export_cue_sheet_to_excel
from backend.exporters.cisac_xml import export_cue_sheet_to_cisac_xml

@pytest.fixture
def sample_manifest():
    cues = [
        ResolvedCue(
            cue_number=1,
            title="Midnight City",
            artist="M83",
            usage_type=UsageType.BI,
            timecode_in="00:00:05:00",
            timecode_out="00:01:31:12",
            duration_frames=2076,
            duration_timecode="00:01:26:12",
            work_id="ASCAP-883491024",
            iswc="T-909.123.456-1",
            writers=[
                RightsHolder(name="Anthony Gonzalez", share=50.0, pro="SACEM"),
                RightsHolder(name="Morgan Kibby", share=25.0, pro="BMI"),
                RightsHolder(name="Yann Gonzalez", share=25.0, pro="SACEM")
            ],
            publishers=[
                RightsHolder(name="Delphic Music", share=50.0, pro="ASCAP"),
                RightsHolder(name="Universal Music", share=50.0, pro="BMI")
            ],
            is_verified=True,
            total_writer_share=100.0,
            total_publisher_share=100.0,
            split_status=SplitStatus.CONFIRMED_PUBLIC_SPLIT
        ),
        ResolvedCue(
            cue_number=2,
            title="Time",
            artist="Hans Zimmer",
            usage_type=UsageType.BI,
            timecode_in="00:01:32:00",
            timecode_out="00:02:17:00",
            duration_frames=1080,
            duration_timecode="00:00:45:00",
            work_id="BMI-12849102",
            iswc="T-902.987.654-3",
            writers=[
                RightsHolder(name="Hans Florian Zimmer", share=100.0, pro="BMI")
            ],
            publishers=[
                RightsHolder(name="Warner-Barham Music LLC", share=100.0, pro="BMI")
            ],
            is_verified=True,
            total_writer_share=100.0,
            total_publisher_share=100.0,
            split_status=SplitStatus.CONFIRMED_PUBLIC_SPLIT
        ),
        # Case B: Undisclosed split track (Radiohead) with share=None
        ResolvedCue(
            cue_number=3,
            title="Exit Music (For A Film)",
            artist="Radiohead",
            usage_type=UsageType.BI,
            timecode_in="00:02:18:00",
            timecode_out="00:03:00:00",
            duration_frames=1008,
            duration_timecode="00:00:42:00",
            work_id="BMI-4819203",
            iswc="T-010.456.789-0",
            writers=[
                RightsHolder(name="Thomas Edward Yorke", share=None, pro="BMI"),
                RightsHolder(name="Jonathan Richard Guy Greenwood", share=None, pro="BMI")
            ],
            publishers=[
                RightsHolder(name="Warner Chappell Music Ltd", share=None, pro="BMI")
            ],
            is_verified=False,
            total_writer_share=0.0,
            total_publisher_share=0.0,
            split_status=SplitStatus.PRO_REGISTERED_SPLIT_UNDISCLOSED,
            estimated_equal_share=50.0
        )
    ]
    return CueSheetManifest(
        project_title="Neon Mirage Trailer",
        production_company="Studio Alpha",
        director="Jane Director",
        target_distributor="Netflix",
        cues=cues,
        total_cues=3,
        cleared_cues=2,
        flagged_cues=1,
        compliance_score=66.7
    )

def test_excel_export(sample_manifest):
    # Guarantees that Case B undisclosed cues with share=None do not raise TypeError
    excel_bytes = export_cue_sheet_to_excel(sample_manifest)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 2000

    # Load with openpyxl to verify integrity
    wb = load_workbook(io.BytesIO(excel_bytes))
    assert "Music Cue Sheet" in wb.sheetnames
    
    ws = wb["Music Cue Sheet"]
    assert "NEON MIRAGE TRAILER" in ws["A1"].value
    assert ws["B9"].value == "Midnight City"
    assert ws["B10"].value == "Time"
    assert ws["B11"].value == "Exit Music (For A Film)"
    
    # Assert Status formatting
    assert ws["M9"].value == "CLEARED"
    assert ws["M10"].value == "CLEARED"
    assert "ACTION REQUIRED" in ws["M11"].value

    # Assert that undisclosed shares are rendered as "Undisclosed" without crashing
    assert "Undisclosed" in ws["I11"].value
    assert "Undisclosed" in ws["K11"].value

def test_cisac_xml_export(sample_manifest):
    # Guarantees that Case B undisclosed cues with share=None do not raise TypeError
    xml_content = export_cue_sheet_to_cisac_xml(sample_manifest)
    assert isinstance(xml_content, str)
    assert "AVCueSheet" in xml_content

    # Parse and verify DOM structure
    root = ET.fromstring(xml_content)
    assert root.tag.endswith("AVCueSheet")
    
    # Assert standard CISAC nodes
    assert "<Sender>CueClear AI Autonomous Clearance Engine</Sender>" in xml_content
    assert "<WorkOrigin>Film/TV Post-Production Sequence</WorkOrigin>" in xml_content
    assert "<WorkTitle>Neon Mirage Trailer</WorkTitle>" in xml_content
    assert "Midnight City" in xml_content
    assert "Time" in xml_content
    assert "Exit Music (For A Film)" in xml_content

    # Assert that UNDISCLOSED shares are rendered properly in XML
    assert "<Share>UNDISCLOSED</Share>" in xml_content
