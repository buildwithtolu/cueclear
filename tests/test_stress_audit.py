import pytest
import io
import asyncio
import xml.etree.ElementTree as ET
from unittest.mock import patch, AsyncMock
from openpyxl import load_workbook

from backend.parsers.edl_parser import EDLParser
from backend.parsers.xml_parser import XMLParser
from backend.agent.parallel_tool import clean_music_search_terms, search_pro_music_rights
from backend.models.schemas import (
    ParsedAudioClip,
    ResolvedCue,
    RightsHolder,
    UsageType,
    SplitStatus,
    CueSheetManifest
)
from backend.agent.split_reconciler import validate_splits, compute_manifest_compliance
from backend.agent.adk_orchestrator import CueClearADKOrchestrator
from backend.exporters.excel_exporter import export_cue_sheet_to_excel
from backend.exporters.cisac_xml import export_cue_sheet_to_cisac_xml

# ==============================================================================
# SCENARIO 1: Arbitrary & Noisy Timeline Ingestion (Parser Resilience)
# ==============================================================================

def test_scenario_1_arbitrary_and_noisy_timeline_ingestion():
    """
    Verifies that EDL and XML parsers:
    1. Filter out video cuts (V, V1, VA/V).
    2. Deduplicate linked stereo channels (A1 & A2 with matching timecodes).
    3. Normalize noisy stems (sample rates, channel specs, version tags) without crashing.
    """
    noisy_edl = """
    TITLE: NOISY_MASTER_TIMELINE_V3
    FCM: NON-DROP FRAME

    001  V        C        00:00:00:00 00:00:10:00 01:00:00:00 01:00:10:00
    * FROM CLIP NAME: 4K_Broll_DroneShot_v2.mov

    002  VA/V     C        00:00:10:00 00:00:15:00 01:00:10:00 01:00:15:00
    * FROM CLIP NAME: Interview_A_Cam.mp4

    003  A1       C        00:00:00:00 00:01:00:00 01:00:15:00 01:01:15:00
    * FROM CLIP NAME: 01_M83_Midnight_City_48k_Stereo_v3_final.wav

    004  A2       C        00:00:00:00 00:01:00:00 01:00:15:00 01:01:15:00
    * FROM CLIP NAME: 01_M83_Midnight_City_48k_Stereo_v3_final.wav

    005  AUD      C        00:00:00:00 00:00:05:00 01:01:15:00 01:01:20:00
    * FROM CLIP NAME: SFX_Gunshot_Stereo_96k_master.wav
    """
    parser = EDLParser(default_fps=24.0)
    clips = parser.parse(noisy_edl)

    # 1. Assert video cuts (001, 002) were skipped
    assert len(clips) == 2, f"Expected 2 audio events after video filter & stereo dedup, got {len(clips)}"
    
    # 2. Assert stereo linked tracks (003 & 004) were deduplicated to 1 cue
    assert "Midnight City" in clips[0].clip_name
    assert clips[0].record_in == "01:00:15:00"
    assert clips[0].record_out == "01:01:15:00"

    # 3. Assert clean_music_search_terms normalization
    title1, artist1 = clean_music_search_terms(clips[0].clip_name)
    assert title1 == "M83 Midnight City"
    assert "48k" not in title1.lower()
    assert "stereo" not in title1.lower()
    assert "v3" not in title1.lower()
    assert "final" not in title1.lower()

    title2, _ = clean_music_search_terms("Unknown_Demo_Track_96k_mono_ver2.mp3")
    assert title2 == "Unknown Demo Track"


# ==============================================================================
# SCENARIO 2: Stream Fault Isolation & Rate-Limit Retries (SSE Resilience)
# ==============================================================================

@pytest.mark.asyncio
async def test_scenario_2_stream_fault_isolation_and_rate_limit_retries():
    """
    Simulates a network timeout / 429 failure during a multi-cue timeline stream.
    Verifies that:
    1. The agent catches the isolated exception.
    2. Logs a warning to the SSE stream.
    3. Continues processing subsequent cues to completion without crashing the server.
    """
    clips = [
        ParsedAudioClip(
            event_number=1, track_type="A1", clip_name="Studio Ident Logo",
            source_in="00:00:00:00", source_out="00:00:05:00",
            record_in="01:00:00:00", record_out="01:00:05:00",
            duration_frames=120, duration_timecode="00:00:05:00"
        ),
        ParsedAudioClip(
            event_number=2, track_type="A1", clip_name="Corrupted_Network_Track",
            source_in="00:00:00:00", source_out="00:01:00:00",
            record_in="01:00:05:00", record_out="01:01:05:00",
            duration_frames=1440, duration_timecode="00:01:00:00"
        ),
        ParsedAudioClip(
            event_number=3, track_type="A1", clip_name="Hans Zimmer - Time",
            source_in="00:00:00:00", source_out="00:00:45:00",
            record_in="01:01:05:00", record_out="01:01:50:00",
            duration_frames=1080, duration_timecode="00:00:45:00"
        )
    ]

    agent = CueClearADKOrchestrator()
    events = []

    # Mock parallel tool to raise an unexpected runtime network error specifically on Cue 2
    original_search = agent.runner.agent.tools[0] if agent.runner and agent.runner.agent else None

    async def mock_parallel_with_fault(title: str, artist: str = ""):
        if "Corrupted" in title:
            raise ConnectionResetError("Simulated HTTP 429 / Remote Connection Reset")
        return await search_pro_music_rights(title, artist)

    with patch("backend.agent.adk_orchestrator.parallel_pro_search_tool", side_effect=mock_parallel_with_fault):
        async for event in agent.process_timeline_stream(clips, "Fault Isolation Reel"):
            events.append(event)

    # Assert stream completed with start and complete events
    assert len(events) > 0
    assert events[0].event_type == "start"
    assert events[-1].event_type == "complete"

    manifest_data = events[-1].data
    assert manifest_data is not None
    assert manifest_data["total_cues"] == 3
    # Cue 1 (In-House SFX) cleared, Cue 2 (Fault fallback) flagged, Cue 3 processed
    assert manifest_data["cleared_cues"] >= 1
    assert any("Notice: Isolated exception on cue #2" in e.message for e in events if e.event_type == "reasoning")


# ==============================================================================
# SCENARIO 3: Domain Split Precision (Case A, Case B, Case C)
# ==============================================================================

def test_scenario_3_domain_split_precision_cases():
    """
    Tests rigorous 3-tier domain split classification:
    - Case A: Confirmed 100%/100% splits -> Cleared (100% compliance numerator)
    - Case B: PRO registered undisclosed splits -> 0% compliance numerator + equal share estimate
    - Case C: Incomplete publisher claim (<100%) -> 0% compliance numerator + flagged alert
    """
    # Case A: Midnight City (Confirmed splits)
    cue_a = ResolvedCue(
        cue_number=1, title="Midnight City", artist="M83", usage_type=UsageType.BI,
        timecode_in="00:00:00:00", timecode_out="00:01:00:00", duration_frames=1440, duration_timecode="00:01:00:00",
        writers=[
            RightsHolder(name="Anthony Gonzalez", share=50.0, pro="SACEM"),
            RightsHolder(name="Morgan Kibby", share=25.0, pro="BMI"),
            RightsHolder(name="Yann Gonzalez", share=25.0, pro="SACEM")
        ],
        publishers=[
            RightsHolder(name="Delphic Music", share=50.0, pro="ASCAP"),
            RightsHolder(name="Universal Music", share=50.0, pro="BMI")
        ]
    )
    is_a, flags_a = validate_splits(cue_a)
    assert is_a is True
    assert cue_a.split_status == SplitStatus.CONFIRMED_PUBLIC_SPLIT

    # Case B: Exit Music (Undisclosed splits)
    cue_b = ResolvedCue(
        cue_number=2, title="Exit Music (For A Film)", artist="Radiohead", usage_type=UsageType.BI,
        timecode_in="00:01:00:00", timecode_out="00:02:00:00", duration_frames=1440, duration_timecode="00:01:00:00",
        writers=[
            RightsHolder(name="Thom Yorke", share=None, pro="BMI"),
            RightsHolder(name="Jonny Greenwood", share=None, pro="BMI"),
            RightsHolder(name="Colin Greenwood", share=None, pro="BMI"),
            RightsHolder(name="Ed O'Brien", share=None, pro="BMI"),
            RightsHolder(name="Philip Selway", share=None, pro="BMI")
        ],
        publishers=[
            RightsHolder(name="Warner Chappell", share=None, pro="BMI")
        ]
    )
    is_b, flags_b = validate_splits(cue_b)
    assert is_b is False
    assert cue_b.split_status == SplitStatus.PRO_REGISTERED_SPLIT_UNDISCLOSED
    assert cue_b.estimated_equal_share == 20.0  # 100 / 5 writers

    # Case C: Partial Publisher Claim (75% claimed, 25% unrepresented)
    cue_c = ResolvedCue(
        cue_number=3, title="Indie Collaboration", artist="Indie Band", usage_type=UsageType.BI,
        timecode_in="00:02:00:00", timecode_out="00:03:00:00", duration_frames=1440, duration_timecode="00:01:00:00",
        writers=[
            RightsHolder(name="Composer A", share=100.0, pro="ASCAP")
        ],
        publishers=[
            RightsHolder(name="Publisher A", share=75.0, pro="ASCAP")  # Missing 25%
        ]
    )
    is_c, flags_c = validate_splits(cue_c)
    assert is_c is False
    assert cue_c.split_status == SplitStatus.PARTIAL_PUBLISHER_CLAIM_FLAGGED

    # Manifest Scorer: 1 cleared out of 3 -> exactly 33.3%
    manifest = compute_manifest_compliance([cue_a, cue_b, cue_c], "3-Tier Compliance Reel")
    assert manifest.total_cues == 3
    assert manifest.cleared_cues == 1
    assert manifest.flagged_cues == 2
    assert manifest.compliance_score == 33.3


# ==============================================================================
# SCENARIO 4: Export Deliverable Resilience (Excel & CISAC XML)
# ==============================================================================

def test_scenario_4_export_deliverable_resilience_mixed_manifest():
    """
    Verifies that exporting a manifest with mixed Case A, Case B (share=None), and Case C cues:
    1. Generates valid Excel and XML bytes without TypeError.
    2. Correctly renders "Undisclosed" in Excel share cells and "ACTION REQUIRED".
    3. Correctly renders "<Share>UNDISCLOSED</Share>" and standard CISAC nodes in XML.
    """
    cue_a = ResolvedCue(
        cue_number=1, title="Midnight City", artist="M83", usage_type=UsageType.BI,
        timecode_in="00:00:00:00", timecode_out="00:01:00:00", duration_frames=1440, duration_timecode="00:01:00:00",
        writers=[RightsHolder(name="Anthony Gonzalez", share=100.0, pro="SACEM")],
        publishers=[RightsHolder(name="Delphic Music", share=100.0, pro="ASCAP")],
        is_verified=True, split_status=SplitStatus.CONFIRMED_PUBLIC_SPLIT
    )
    cue_b = ResolvedCue(
        cue_number=2, title="Exit Music", artist="Radiohead", usage_type=UsageType.BI,
        timecode_in="00:01:00:00", timecode_out="00:02:00:00", duration_frames=1440, duration_timecode="00:01:00:00",
        writers=[RightsHolder(name="Thom Yorke", share=None, pro="BMI")],
        publishers=[RightsHolder(name="Warner Chappell", share=None, pro="BMI")],
        is_verified=False, split_status=SplitStatus.PRO_REGISTERED_SPLIT_UNDISCLOSED,
        estimated_equal_share=100.0
    )
    cue_c = ResolvedCue(
        cue_number=3, title="Partial Track", artist="Indie", usage_type=UsageType.BI,
        timecode_in="00:02:00:00", timecode_out="00:03:00:00", duration_frames=1440, duration_timecode="00:01:00:00",
        writers=[RightsHolder(name="Composer A", share=100.0, pro="ASCAP")],
        publishers=[RightsHolder(name="Publisher A", share=60.0, pro="BMI")],
        is_verified=False, split_status=SplitStatus.PARTIAL_PUBLISHER_CLAIM_FLAGGED
    )

    manifest = CueSheetManifest(
        project_title="Stress Test Reel",
        production_company="Universal Post",
        director="Chief Director",
        target_distributor="Broadcast Global",
        cues=[cue_a, cue_b, cue_c],
        total_cues=3, cleared_cues=1, flagged_cues=2, compliance_score=33.3
    )

    # 1. Excel Generation Check
    excel_bytes = export_cue_sheet_to_excel(manifest)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 2000

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb["Music Cue Sheet"]
    assert ws["B9"].value == "Midnight City"
    assert ws["M9"].value == "CLEARED"
    assert ws["B10"].value == "Exit Music"
    assert "ACTION REQUIRED" in ws["M10"].value
    assert "Undisclosed" in ws["I10"].value

    # 2. CISAC XML Generation Check
    xml_str = export_cue_sheet_to_cisac_xml(manifest)
    assert isinstance(xml_str, str)
    assert "<Sender>CueClear AI Autonomous Clearance Engine</Sender>" in xml_str
    assert "<WorkOrigin>Film/TV Post-Production Sequence</WorkOrigin>" in xml_str
    assert "<WorkTitle>Stress Test Reel</WorkTitle>" in xml_str
    assert "<Share>UNDISCLOSED</Share>" in xml_str
